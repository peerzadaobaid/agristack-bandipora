"""
AGRISTACK Dashboard — District Bandipora

Tabs (left to right):
  1. Tehsil Wise      — S.No | Tehsil | Villages | Total Survey Nos | Submitted | % Completion | Additions | % Approved
  2. Overall All      — S.No | Tehsil | Total Khasras | Khasras Submitted | Khasras Verified | Khasras Approved
  3. Patwari Wise     — By % Completion (with Relative Effort) and By Total Submissions
  4. Checker Wise     — S.No | Checker | Sub-Division | Villages | Total Survey Nos | Submitted | Verified+Approved | % Completion | Additions
                       (% Completion here = (V + A + Seek Clarification) / Submitted)
  5. Village Wise     — S.No | Tehsil | Village | Total Survey Nos | Name of Patwari | Submitted | Verified | Approved | Additions

Sub-Division is derived from Tehsil via a hardcoded map (robust to broken SUB DIVISION source columns).
Tehsil filter checkboxes apply to Patwari and Village views: filter + high→low sort + color banding.
"""
import os
import re
import io
import sys
from datetime import datetime, date

import pandas as pd
from flask import Flask, render_template, request, Response, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from xhtml2pdf import pisa

app = Flask(__name__)

SNAPSHOTS_DIR = "snapshots"
LEGACY_EXCEL = "AGRISTACK.xlsx"
REFERENCE_PATH = "reference.xlsx"

# Target Based view configuration
DEADLINE_DATE = date(2026, 9, 30)          # District-wide plan deadline
PLAN_START_DATE = date(2026, 9, 5)         # When patwaris begin executing the plan
BASELINE_TARGET_DATE = date(2026, 9, 4)    # Snapshot used as baseline for rate calculation
FIVE_DAY_WINDOWS = [
    ("06-10 Sep", date(2026, 9, 6),  date(2026, 9, 10)),
    ("11-15 Sep", date(2026, 9, 11), date(2026, 9, 15)),
    ("16-20 Sep", date(2026, 9, 16), date(2026, 9, 20)),
    ("21-25 Sep", date(2026, 9, 21), date(2026, 9, 25)),
    ("26-30 Sep", date(2026, 9, 26), date(2026, 9, 30)),
]

# Rate-ratio bands (Green >= 80%, Yellow 50-79%, Red < 50%)
GREEN_MIN = 80.0
YELLOW_MIN = 50.0

# Recognize dated snapshot filenames in two shapes:
#   1. Our own: "2026-08-13.xlsx"     (YYYY-MM-DD)
#   2. Government REPORT: "REPORT 21.08.2026.xlsx" or "REPORT_21_08_2026.xlsx"
#      or "REPORT_21_08_2026__1_.xlsx" (as the file downloads with a suffix)
_DATE_FILENAME_PATTERNS = [
    (re.compile(r"^(\d{4})-(\d{2})-(\d{2})\.xlsx$", re.I),
     lambda m: date(int(m.group(1)), int(m.group(2)), int(m.group(3)))),
    (re.compile(r"^REPORT[\s_](\d{1,2})[._](\d{1,2})[._](\d{4}).*\.xlsx$", re.I),
     lambda m: date(int(m.group(3)), int(m.group(2)), int(m.group(1)))),
]


def _parse_snapshot_date(filename):
    """Return the date embedded in a snapshot filename, or None."""
    for rx, extractor in _DATE_FILENAME_PATTERNS:
        m = rx.match(filename)
        if m:
            try:
                return extractor(m)
            except ValueError:
                return None
    return None

# Tehsil -> Sub-Division (stable mapping; used because the SUB DIVISION column in source
# files has been unreliable — sometimes text, sometimes numbers). If any tehsil hierarchy
# ever changes, edit here.
TEHSIL_TO_SUBDIV = {
    "BANDIPORA": "BANDIPORA",
    "ALOOSA":    "BANDIPORA",
    "AJAS":      "BANDIPORA",
    "SUMBAL":    "SUMBAL",
    "HAJIN":     "SUMBAL",
    "GUREZ":     "GUREZ",
    "TULAIL":    "GUREZ",
}

COL_MATCHERS = {
    "tehsil":    re.compile(r"tehsil(?!dar)", re.I),   # not TEHSILDAR
    "village":   re.compile(r"^village$", re.I),
    "patwari":   re.compile(r"patwari", re.I),
    "khasras":   re.compile(r"khasra|survey", re.I),
    # 'submitted' is the patwari's total output. The new file introduces a "Total Submitted"
    # column which represents this (the file's "submitted" column now means "in-queue awaiting
    # first check" — a workflow stage, not the total). Prefer "Total Submitted" when present;
    # fall back to plain "submitted" for older snapshot files.
    "submitted": re.compile(r"^(total\s+)?submitted$", re.I),
}

# Optional workflow columns — added recently. Loaded when present, treated as 0 otherwise.
OPTIONAL_MATCHERS = {
    "checker":            re.compile(r"maker|checker", re.I),        # CONCERNED MAKER = checker
    "approved":           re.compile(r"^approved$", re.I),           # Approved
    "verified":           re.compile(r"^verified$", re.I),           # verified (excludes approved rows)
    "seek_clarification": re.compile(r"seek.*clar|clarif", re.I),    # Seek Clarification
}


# ---------- Snapshot discovery ----------

def all_snapshots():
    """Return [(date, path), ...] newest first. Accepts both YYYY-MM-DD.xlsx
    and REPORT DD.MM.YYYY.xlsx filename styles. Ignores plan_YYYY-MM-DD.xlsx
    (those are handled by load_plan_file)."""
    out = []
    if os.path.isdir(SNAPSHOTS_DIR):
        for name in os.listdir(SNAPSHOTS_DIR):
            if not name.lower().endswith(".xlsx") or name.startswith("~"):
                continue
            if name.lower().startswith("plan_") or name.lower().startswith("plan "):
                continue
            d = _parse_snapshot_date(name)
            if d:
                out.append((d, os.path.join(SNAPSHOTS_DIR, name)))
    if not out and os.path.exists(LEGACY_EXCEL):
        mtime = os.path.getmtime(LEGACY_EXCEL)
        out.append((datetime.fromtimestamp(mtime).date(), LEGACY_EXCEL))
    out.sort(key=lambda x: x[0], reverse=True)
    return out


def snapshot_for_date(snapshots, target_date):
    for d, p in snapshots:
        if d == target_date:
            return p
    return None


# ---------- Excel loading ----------

_df_cache = {}


def _pick_sheet(xl):
    """Prefer Sheet2 (corrected), then MAIN, then first sheet scoring highest on required columns."""
    scored = []
    for name in xl.sheet_names:
        try:
            head = pd.read_excel(xl, sheet_name=name, nrows=0)
        except Exception:
            continue
        keys = [str(c) for c in head.columns]
        score = 0
        for logical, rx in COL_MATCHERS.items():
            for k in keys:
                if rx.search(k.strip()):
                    if logical == "khasras" and re.search(r"submitted", k, re.I):
                        continue
                    score += 1
                    break
        rank = 0 if name == "Sheet2" else 1 if name == "MAIN" else 2
        scored.append((score, -rank, name))
    scored.sort(reverse=True)
    if not scored:
        raise RuntimeError("No usable sheets found.")
    picked = scored[0][2]
    print(f"[sheet-picker] Picked '{picked}' from {[n for _, _, n in scored]}", file=sys.stderr)
    return picked


def _detect_columns(df):
    cols = {}
    for logical, rx in COL_MATCHERS.items():
        for c in df.columns:
            key = str(c).strip()
            if rx.search(key):
                if logical == "khasras" and re.search(r"submitted", key, re.I):
                    continue
                cols[logical] = c
                break
    missing = [k for k in COL_MATCHERS if k not in cols]
    if missing:
        raise RuntimeError(f"Missing required columns: {', '.join(missing)}")
    for logical, rx in OPTIONAL_MATCHERS.items():
        for c in df.columns:
            key = str(c).strip()
            if rx.search(key):
                cols[logical] = c
                break
    return cols


def _is_report_format(df_columns):
    """Detect the government REPORT format: has 'Total' + workflow stages but
    lacks patwari/checker/TOTAL KHASRAS columns."""
    cols_lower = [str(c).strip().lower() for c in df_columns]
    has_report_total = "total" in cols_lower and not any("total khasra" in c or "total survey" in c for c in cols_lower)
    has_report_stages = "submitted" in cols_lower and "verified" in cols_lower and "approved" in cols_lower
    has_patwari = any("patwari" in c for c in cols_lower)
    has_workload = any("total khasra" in c or "total survey" in c for c in cols_lower)
    return has_report_total and has_report_stages and not has_patwari and not has_workload


_reference_cache = {}
_plan_cache = {}


def has_plan_file():
    """Cheap filesystem-only check: is any plan_*.xlsx present? Does NOT read
    the file. Use this to decide whether to show the Target Based view toggle."""
    if not os.path.isdir(SNAPSHOTS_DIR):
        return False
    for name in os.listdir(SNAPSHOTS_DIR):
        if re.match(r"plan[_ ]\d{4}-\d{2}-\d{2}\.xlsx$", name, re.I):
            return True
    return False


def load_plan_file():
    """Find and load the newest plan_YYYY-MM-DD.xlsx in snapshots/. Returns a
    DataFrame with (tehsil, village, expected_date, is_completed_per_plan,
    start_order) or None if no plan file present."""
    if not os.path.isdir(SNAPSHOTS_DIR):
        return None
    plan_files = []
    for name in os.listdir(SNAPSHOTS_DIR):
        m = re.match(r"plan[_ ](\d{4}-\d{2}-\d{2})\.xlsx$", name, re.I)
        if not m:
            continue
        try:
            d = datetime.strptime(m.group(1), "%Y-%m-%d").date()
        except ValueError:
            continue
        plan_files.append((d, os.path.join(SNAPSHOTS_DIR, name)))
    if not plan_files:
        return None
    plan_files.sort(key=lambda x: x[0], reverse=True)
    plan_date, plan_path = plan_files[0]
    mtime = os.path.getmtime(plan_path)
    cache_key = (plan_path, mtime)
    if cache_key in _plan_cache:
        return _plan_cache[cache_key]
    # Sheet 1 has a title, rules row, colour-key row, then the header on row index 3.
    raw = pd.read_excel(plan_path, sheet_name="Plan by Patwari", header=3)
    # Keep only rows with a numeric S.No and non-empty Village (drops TOTAL & Notes)
    raw = raw[raw["Village"].notna() & (raw["Village"].astype(str).str.strip() != "")]
    raw = raw[raw["S.No"].apply(lambda x: isinstance(x, (int, float)) and pd.notna(x))]
    exp_str = raw["Expected Date of Completion"].astype(str).str.strip()
    is_completed = exp_str.str.lower() == "completed"
    exp_date = pd.to_datetime(raw["Expected Date of Completion"], errors="coerce").dt.date
    result = pd.DataFrame({
        "tehsil_key": raw["Tehsil"].astype(str).str.strip().str.upper().values,
        "village_key": raw["Village"].astype(str).str.strip().values,
        "expected_date": exp_date.values,
        "is_completed_per_plan": is_completed.values,
        "start_order": raw["Start Order"].astype(str).str.strip().values,
    })
    print(f"[plan] Loaded {len(result)} rows from {plan_path} (dated {plan_date})", file=sys.stderr)
    _plan_cache[cache_key] = result
    return result


def find_baseline_snapshot(snapshots, target_date=BASELINE_TARGET_DATE):
    """Choose the snapshot that best serves as the rate-calculation baseline.
    Prefer an exact match on `target_date`. Otherwise, use the snapshot closest
    to `target_date` (measured in absolute days)."""
    if not snapshots:
        return None
    # snapshots is [(date, path), ...] newest first
    for d, p in snapshots:
        if d == target_date:
            return (d, p)
    # Fall back to closest snapshot by absolute day distance
    closest = min(snapshots, key=lambda x: abs((x[0] - target_date).days))
    return closest


def load_reference():
    """Read reference.xlsx and return a DataFrame with static columns:
    tehsil, village, patwari, checker, khasras. Returns None if no reference
    file exists. Cache is keyed on file mtime so live edits invalidate."""
    if not os.path.exists(REFERENCE_PATH):
        return None
    mtime = os.path.getmtime(REFERENCE_PATH)
    cache_key = (REFERENCE_PATH, mtime)
    if cache_key in _reference_cache:
        return _reference_cache[cache_key]
    xl = pd.ExcelFile(REFERENCE_PATH)
    sheet = _pick_sheet(xl)
    df = pd.read_excel(xl, sheet_name=sheet)
    cols = _detect_columns(df)
    # Extract only the static columns needed for merge
    result = pd.DataFrame({
        "tehsil": df[cols["tehsil"]].astype(str).str.strip().str.upper(),
        "village": df[cols["village"]].astype(str).str.strip(),
        "patwari": df[cols["patwari"]].astype(str).str.strip().replace({"nan": "", "None": ""}),
        "khasras": pd.to_numeric(df[cols["khasras"]], errors="coerce").fillna(0).astype(int),
        "checker": df[cols["checker"]].astype(str).str.strip().replace({"nan": "", "None": ""}) if "checker" in cols else "",
    })
    result = result[(result["village"] != "") & (result["village"].str.lower() != "nan")]
    print(f"[reference] Loaded {len(result)} villages from {REFERENCE_PATH}", file=sys.stderr)
    _reference_cache[cache_key] = result
    return result


def _load_report_and_merge(path):
    """Load a REPORT-format file, merge with reference, and return the
    standardized DataFrame."""
    ref = load_reference()
    if ref is None:
        raise RuntimeError(
            f"Snapshot '{os.path.basename(path)}' is in REPORT format (no patwari/khasras "
            f"columns) but no reference.xlsx was found in the repo root. Please upload "
            f"reference.xlsx containing tehsil, village, patwari, checker, and TOTAL KHASRAS."
        )
    xl = pd.ExcelFile(path)
    sheet = _pick_sheet(xl)
    raw = pd.read_excel(xl, sheet_name=sheet)
    # REPORT format has a blank first row (all NaN) — drop it and any similar
    key_cols = [c for c in raw.columns if str(c).strip().lower() in ("tehsil", "village")]
    if key_cols:
        raw = raw.dropna(subset=key_cols)

    # Column lookup, case-insensitive
    def find(names):
        for c in raw.columns:
            if str(c).strip().lower() in [n.lower() for n in names]:
                return c
        return None
    c_tehsil = find(["tehsil"])
    c_village = find(["village"])
    c_submitted = find(["submitted"])                 # workflow stage (in-queue)
    c_seek = find(["seek clarification"])
    c_resub = find(["re submitted", "resubmitted"])
    c_verified = find(["verified"])
    c_approved = find(["approved"])

    if not all([c_tehsil, c_village, c_verified, c_approved]):
        raise RuntimeError("REPORT file is missing required columns (Tehsil/Village/Verified/Approved).")

    # Compute Total Submitted = sum of workflow stages (this is what our dashboard calls "submitted")
    def _num(col):
        return pd.to_numeric(raw[col], errors="coerce").fillna(0).astype(int) if col else pd.Series(0, index=raw.index, dtype=int)

    stage_submitted = _num(c_submitted)
    stage_seek = _num(c_seek)
    stage_resub = _num(c_resub)
    stage_verified = _num(c_verified)
    stage_approved = _num(c_approved)
    total_submitted = stage_submitted + stage_seek + stage_resub + stage_verified + stage_approved

    rpt = pd.DataFrame({
        "tehsil_key": raw[c_tehsil].astype(str).str.strip().str.upper(),
        "village_key": raw[c_village].astype(str).str.strip(),
        "submitted": total_submitted,
        "seek_clarification": stage_seek,
        "verified": stage_verified,
        "approved": stage_approved,
    })

    # Strict merge on (tehsil_key, village_key)
    ref_keyed = ref.copy()
    ref_keyed["tehsil_key"] = ref_keyed["tehsil"]  # already uppercased in load_reference
    ref_keyed["village_key"] = ref_keyed["village"]
    merged = rpt.merge(
        ref_keyed[["tehsil_key", "village_key", "tehsil", "village", "patwari", "checker", "khasras"]],
        on=["tehsil_key", "village_key"],
        how="left",
        indicator=True,
    )
    unmatched_rpt = merged[merged["_merge"] == "left_only"]
    if len(unmatched_rpt):
        rows = [f"  - {r['tehsil_key']} / {r['village_key']}" for _, r in unmatched_rpt.head(20).iterrows()]
        print(f"[reference-merge] WARNING: {len(unmatched_rpt)} REPORT rows have no match in reference:\n"
              + "\n".join(rows), file=sys.stderr)
    # Also check for reference villages missing from REPORT (optional info)
    ref_only = ref_keyed.merge(
        rpt[["tehsil_key", "village_key"]], on=["tehsil_key", "village_key"], how="left", indicator=True
    )
    ref_only = ref_only[ref_only["_merge"] == "left_only"]
    if len(ref_only):
        print(f"[reference-merge] Note: {len(ref_only)} villages in reference not present in this REPORT "
              f"(will show as 0 activity).", file=sys.stderr)

    # Keep only matched rows for the dashboard; use REPORT numbers + reference static info
    ok = merged[merged["_merge"] == "both"].copy()
    result = pd.DataFrame({
        "tehsil": ok["tehsil"],   # canonical casing from reference
        "village": ok["village"],
        "patwari": ok["patwari"],
        "khasras": ok["khasras"],
        "submitted": ok["submitted"],
        "checker": ok["checker"],
        "approved": ok["approved"],
        "verified": ok["verified"],
        "seek_clarification": ok["seek_clarification"],
    })
    return result


def load_snapshot(path):
    if path in _df_cache:
        return _df_cache[path]

    # Peek at columns to decide format
    xl = pd.ExcelFile(path)
    sheet = _pick_sheet(xl)
    peek = pd.read_excel(xl, sheet_name=sheet, nrows=1)
    if _is_report_format(peek.columns):
        # REPORT format — merge with reference
        new_df = _load_report_and_merge(path)
    else:
        # Full format — existing behavior
        df = pd.read_excel(xl, sheet_name=sheet)
        cols = _detect_columns(df)
        new_df = pd.DataFrame()
        for logical in ("tehsil", "village", "patwari", "khasras", "submitted"):
            new_df[logical] = df[cols[logical]].values
        for opt in ("checker", "approved", "verified", "seek_clarification"):
            if opt in cols:
                new_df[opt] = df[cols[opt]].values

    # Numeric coercion + string cleanup (common to both formats)
    # Explicit .copy() to eliminate SettingWithCopyWarning / ChainedAssignmentError.
    new_df = new_df.copy()
    for numcol in ("khasras", "submitted", "approved", "verified", "seek_clarification"):
        if numcol in new_df.columns:
            new_df[numcol] = pd.to_numeric(new_df[numcol], errors="coerce").fillna(0).astype(int)
        else:
            new_df[numcol] = 0
    for scol in ("tehsil", "village", "patwari", "checker"):
        if scol in new_df.columns:
            new_df[scol] = new_df[scol].astype(str).str.strip().replace({"nan": "", "NaN": "", "None": ""})
    new_df = new_df[(new_df["village"] != "") & (new_df["village"].str.lower() != "nan")].copy()
    new_df["subdivision"] = new_df["tehsil"].str.upper().map(TEHSIL_TO_SUBDIV).fillna("UNKNOWN")
    _df_cache[path] = new_df
    return new_df


# ---------- Helpers ----------

def _pct(num, denom):
    return (num / denom * 100) if denom > 0 else 0.0


def apply_bands(rows, sort_key, ascending=False):
    """Attach 'band' key ('green'|'yellow'|'red') to each row.
    Semantic: green = doing well, red = struggling.
    - ascending=False (default): rows sorted high-to-low by sort_key (best first).
      Top 30% get green, bottom 30% red.
    - ascending=True: rows sorted low-to-high (worst first).
      Top 30% get red (they're the strugglers), bottom 30% green."""
    n = len(rows)
    if n == 0:
        return
    top_n = round(n * 0.3)
    bot_n = round(n * 0.3)
    if ascending:
        top_band, bot_band = "red", "green"
    else:
        top_band, bot_band = "green", "red"
    for i, r in enumerate(rows):
        if i < top_n:
            r["band"] = top_band
        elif i >= n - bot_n:
            r["band"] = bot_band
        else:
            r["band"] = "yellow"


# ---------- View building ----------

def _band_for_ratio(ratio_pct):
    """Return color band based on progress rate ratio (percent)."""
    if ratio_pct is None:
        return None
    if ratio_pct >= GREEN_MIN:
        return "green"
    if ratio_pct >= YELLOW_MIN:
        return "yellow"
    return "red"


def _window_for_date(d):
    """Return the label of the 5-day window containing date d, or None if outside."""
    if d is None:
        return None
    for label, start, end in FIVE_DAY_WINDOWS:
        if start <= d <= end:
            return label
    # Dates before the first window fall into the first bucket
    if d < FIVE_DAY_WINDOWS[0][1]:
        return FIVE_DAY_WINDOWS[0][0]
    # Dates past the last window fall into the last bucket
    return FIVE_DAY_WINDOWS[-1][0]


def build_target_view(current_df, baseline_df, plan_df, today=None):
    """Build the Target Based view data. Requires a plan file. Returns a dict
    with village_rows, tehsil_rows, window_columns, and metadata. Returns None
    if plan_df is missing."""
    if plan_df is None:
        return None
    if today is None:
        today = date.today()

    # Merge plan onto current data (strict on tehsil_key, village_key)
    cur = current_df[["tehsil", "village", "patwari", "khasras", "submitted", "verified", "approved"]].copy()
    cur["tehsil_key"] = cur["tehsil"].str.upper()
    cur["village_key"] = cur["village"]
    merged = cur.merge(plan_df, on=["tehsil_key", "village_key"], how="left")

    # Baseline: submitted per village at baseline snapshot date, keyed same way
    baseline_map = {}
    baseline_date = None
    if baseline_df is not None:
        bl = baseline_df[["tehsil", "village", "submitted"]].copy()
        bl["k"] = bl["tehsil"].str.upper() + "|" + bl["village"]
        baseline_map = dict(zip(bl["k"], bl["submitted"].astype(int)))

    # For rate: days between baseline and today (must be > 0)
    # In the app, `baseline_date` is passed via meta; here we assume caller
    # supplies baseline_df aligned to BASELINE_TARGET_DATE (or closest).

    village_rows = []
    for _, r in merged.iterrows():
        village = r["village"]
        tehsil = r["tehsil"]
        total = int(r["khasras"])
        submitted = int(r["submitted"])
        verified = int(r["verified"])
        approved = int(r["approved"])
        patwari = r["patwari"]
        expected = r["expected_date"] if pd.notna(r.get("expected_date")) else None
        completed_per_plan = bool(r.get("is_completed_per_plan", False))

        # A village is "actually completed" if approved >= total (strict, per user)
        is_completed_now = (approved >= total and total > 0)

        # Days baseline→today (use max 1 to avoid div by zero)
        baseline_submitted = baseline_map.get(f"{tehsil.upper()}|{village}", 0)

        # Rate calc — only for pending villages with a real expected date
        rate_ratio = None
        band = None
        if not is_completed_now and expected is not None:
            days_to_expected = (expected - today).days
            if days_to_expected <= 0:
                # Past deadline — no meaningful rate; village is delayed
                rate_ratio = 0.0
                band = "red"
            else:
                remaining = max(total - submitted, 0)
                required_daily = remaining / days_to_expected if days_to_expected > 0 else 0
                # actual daily rate is submissions per day since baseline
                # We'll fold `days_since_baseline` in via the caller's supplied value
                # For simplicity, pass through via meta
                rate_ratio = None  # to be filled below

        # Determine pending stage if past expected date and not complete
        pending_stage = None
        days_delayed = None
        if expected is not None and today > expected and not is_completed_now:
            days_delayed = (today - expected).days
            # Determine latest reached stage
            if approved > 0:
                pending_stage = "Approved (partial)"
            elif verified > 0:
                pending_stage = "Verified"
            elif submitted > 0:
                pending_stage = "Submitted"
            else:
                pending_stage = "Not Started"

        village_rows.append({
            "tehsil": tehsil,
            "village": village,
            "patwari": patwari,
            "total": total,
            "submitted": submitted,
            "verified": verified,
            "approved": approved,
            "expected_date": expected,
            "expected_date_str": expected.strftime("%d %b %Y") if expected else "—",
            "is_completed_now": is_completed_now,
            "completed_per_plan": completed_per_plan,
            "pending_stage": pending_stage,
            "days_delayed": days_delayed,
            "baseline_submitted": baseline_submitted,
        })

    # Compute rate ratios using baseline delta
    # days_since_baseline is district-wide (we picked one baseline date)
    days_since_baseline = 0
    if baseline_df is not None:
        # We estimate baseline date from meta if callers pass it; default: 04-Sep
        days_since_baseline = max((today - BASELINE_TARGET_DATE).days, 1)
    for row in village_rows:
        if row["is_completed_now"] or row["expected_date"] is None:
            row["rate_ratio"] = None
            row["band"] = None
            continue
        days_to_expected = (row["expected_date"] - today).days
        if days_to_expected <= 0:
            # Past deadline; treat as extremely under-rate
            row["rate_ratio"] = 0.0
            row["band"] = "red"
            continue
        actual_delta = max(row["submitted"] - row["baseline_submitted"], 0)
        actual_daily = actual_delta / days_since_baseline if days_since_baseline > 0 else 0
        remaining = max(row["total"] - row["submitted"], 0)
        required_daily = remaining / days_to_expected if days_to_expected > 0 else 0
        if required_daily <= 0:
            # Nothing needed — treat as green
            row["rate_ratio"] = 100.0
            row["band"] = "green"
        else:
            ratio = (actual_daily / required_daily) * 100.0
            row["rate_ratio"] = ratio
            row["band"] = _band_for_ratio(ratio)

    # Sort: pending villages by expected date ascending (earliest first, no date = end of pending),
    # then completed villages at the very bottom in tehsil order
    def _sort_key(row):
        if row["is_completed_now"]:
            return (2, "")  # completed section
        if row["expected_date"] is None:
            return (1, row["village"])  # pending but no date
        return (0, row["expected_date"].isoformat())
    village_rows.sort(key=_sort_key)

    # ---------- Tehsil rows ----------
    # Group by tehsil, compute aggregate rates and window counts.
    by_tehsil = {}
    for row in village_rows:
        t = row["tehsil"]
        if t not in by_tehsil:
            by_tehsil[t] = {
                "tehsil": t,
                "villages": 0,
                "total": 0,
                "submitted": 0,
                "baseline_submitted": 0,
                "approved": 0,
                "villages_completed": 0,
                "windows": {label: 0 for label, *_ in FIVE_DAY_WINDOWS},
            }
        e = by_tehsil[t]
        e["villages"] += 1
        e["total"] += row["total"]
        e["submitted"] += row["submitted"]
        e["baseline_submitted"] += row["baseline_submitted"]
        e["approved"] += row["approved"]
        if row["is_completed_now"]:
            e["villages_completed"] += 1
        else:
            # Count in its expected-date window (skips completed villages)
            w = _window_for_date(row["expected_date"])
            if w:
                e["windows"][w] += 1

    # Determine which windows are "pending" (at least one tehsil has non-zero AND today > window end)
    pending_windows = set()
    for label, start, end in FIVE_DAY_WINDOWS:
        if today > end:
            any_pending = any(t["windows"][label] > 0 for t in by_tehsil.values())
            if any_pending:
                pending_windows.add(label)
    # Current or upcoming window: the first window whose end is >= today
    current_window = None
    for label, start, end in FIVE_DAY_WINDOWS:
        if end >= today:
            current_window = label
            break

    # Column definitions for tehsil rows
    window_columns = []
    for label, start, end in FIVE_DAY_WINDOWS:
        if label in pending_windows:
            window_columns.append({"label": label, "header": f"Pending in ({label})", "is_pending": True})
        elif label == current_window:
            window_columns.append({"label": label, "header": f"Next 5-Day Target ({label})", "is_pending": False, "is_current": True})
        # Windows entirely in the past and cleared → dropped
        # Windows entirely in the future (not current) → dropped for now (only "next" shown)

    # Tehsil rate + band
    tehsil_rows = []
    for t_name, e in sorted(by_tehsil.items(), key=lambda x: x[1]["submitted"], reverse=True):
        rate_ratio = None
        band = None
        remaining = max(e["total"] - e["submitted"], 0)
        days_to_deadline = max((DEADLINE_DATE - today).days, 0)
        if days_to_deadline > 0 and remaining > 0:
            required_daily = remaining / days_to_deadline
            actual_delta = max(e["submitted"] - e["baseline_submitted"], 0)
            actual_daily = actual_delta / days_since_baseline if days_since_baseline > 0 else 0
            rate_ratio = (actual_daily / required_daily) * 100.0 if required_daily > 0 else 100.0
            band = _band_for_ratio(rate_ratio)
        elif remaining == 0:
            rate_ratio = 100.0
            band = "green"
        elif days_to_deadline == 0:
            # Deadline reached
            rate_ratio = 0.0
            band = "red"
        # Assemble row with the visible window columns
        window_values = {col["label"]: e["windows"][col["label"]] for col in window_columns}
        tehsil_rows.append({
            "tehsil": t_name,
            "villages": e["villages"],
            "total": e["total"],
            "submitted": e["submitted"],
            "approved": e["approved"],
            "pct": _pct(e["submitted"], e["total"]),
            "pct_approved": _pct(e["approved"], e["total"]),
            "villages_completed": e["villages_completed"],
            "windows": window_values,
            "rate_ratio": rate_ratio,
            "band": band,
        })

    return {
        "village_rows": village_rows,
        "tehsil_rows": tehsil_rows,
        "window_columns": window_columns,
        "meta": {
            "today": today,
            "deadline": DEADLINE_DATE,
            "baseline_target_date": BASELINE_TARGET_DATE,
            "days_since_baseline": days_since_baseline,
            "days_to_deadline": max((DEADLINE_DATE - today).days, 0),
            "green_min": GREEN_MIN,
            "yellow_min": YELLOW_MIN,
        },
    }


def build_views(current_df, from_df=None, tehsils_filter=None):
    """Build every view. tehsils_filter: None | 'all' | set of tehsil names."""
    apply_coloring = tehsils_filter is not None
    active_tehsils = tehsils_filter if isinstance(tehsils_filter, set) else None
    has_additions = from_df is not None

    # === Tehsil Wise ===
    t = (current_df.groupby("tehsil", as_index=False)
                   .agg(villages=("village", "count"),
                        total=("khasras", "sum"),
                        submitted=("submitted", "sum"),
                        approved=("approved", "sum"))
                   .sort_values("submitted", ascending=False))
    old_by_tehsil = from_df.groupby("tehsil")["submitted"].sum().to_dict() if has_additions else {}
    tehsil_rows = []
    for _, row in t.iterrows():
        additions = int(row["submitted"]) - int(old_by_tehsil.get(row["tehsil"], 0)) if has_additions else None
        tehsil_rows.append({
            "tehsil": row["tehsil"],
            "villages": int(row["villages"]),
            "total": int(row["total"]),
            "submitted": int(row["submitted"]),
            "approved": int(row["approved"]),
            "pct": _pct(int(row["submitted"]), int(row["total"])),
            "pct_approved": _pct(int(row["approved"]), int(row["total"])),
            "additions": additions,
        })

    # === Overall All (new) ===
    # Verified column = verified + approved. Approved column = approved alone.
    o = (current_df.groupby("tehsil", as_index=False)
                   .agg(total=("khasras", "sum"),
                        submitted=("submitted", "sum"),
                        verified=("verified", "sum"),
                        approved=("approved", "sum"))
                   .sort_values("submitted", ascending=False))
    overall_rows = []
    for _, row in o.iterrows():
        overall_rows.append({
            "tehsil": row["tehsil"],
            "total": int(row["total"]),
            "submitted": int(row["submitted"]),
            "verified": int(row["verified"]) + int(row["approved"]),  # roll-in per spec
            "approved": int(row["approved"]),
        })
    overall_totals = {
        "total": sum(r["total"] for r in overall_rows),
        "submitted": sum(r["submitted"] for r in overall_rows),
        "verified": sum(r["verified"] for r in overall_rows),
        "approved": sum(r["approved"] for r in overall_rows),
    }

    # === Patwari Wise (two sorts) — filtered by tehsils picker ===
    if active_tehsils is not None:
        p_source = current_df[current_df["tehsil"].str.upper().isin(active_tehsils)].copy()
        p_old_source = from_df[from_df["tehsil"].str.upper().isin(active_tehsils)].copy() if has_additions else None
    else:
        p_source = current_df.copy()
        p_old_source = from_df.copy() if has_additions else None

    p = (p_source.groupby("patwari", as_index=False)
                 .agg(villages_list=("village", lambda s: ", ".join(sorted(s.tolist()))),
                      tehsils=("tehsil", lambda s: " / ".join(sorted(set(s)))),
                      total=("khasras", "sum"),
                      submitted=("submitted", "sum"))).copy()
    p["pct"] = p.apply(lambda r: _pct(int(r["submitted"]), int(r["total"])), axis=1)
    if has_additions and p_old_source is not None:
        old_by_p = p_old_source.groupby("patwari")["submitted"].sum().to_dict()
        p["additions"] = p["patwari"].map(lambda n: int(old_by_p.get(n, 0)))
        p["additions"] = p["submitted"] - p["additions"]
    else:
        p["additions"] = None

    def _p_records(frame):
        return [{
            "patwari": r["patwari"],
            "tehsil": r["tehsils"],
            "villages_list": r["villages_list"],
            "total": int(r["total"]),
            "submitted": int(r["submitted"]),
            "pct": float(r["pct"]),
            "additions": None if r["additions"] is None else int(r["additions"]),
        } for _, r in frame.iterrows()]

    patwari_by_pct = _p_records(p.sort_values(["pct", "submitted"], ascending=[True, True]))
    patwari_by_count = _p_records(p.sort_values(["submitted", "pct"], ascending=[False, False]))

    # Relative Effort — district-wide top submission count, only for By %
    district_top = int(current_df.groupby("patwari")["submitted"].sum().max() or 0)
    for r in patwari_by_pct:
        r["relative_effort"] = (r["submitted"] / district_top * 100) if district_top > 0 else None

    if apply_coloring:
        # By % is sorted low→high: strugglers at top → red at top, green at bottom
        apply_bands(patwari_by_pct, "pct", ascending=True)
        # By Count is sorted high→low: top performers first → green at top, red at bottom
        apply_bands(patwari_by_count, "submitted", ascending=False)

    # Patwari totals for TOTAL row — always from displayed rows so filtered/unfiltered both work
    if patwari_by_pct:
        p_tot = sum(r["total"] for r in patwari_by_pct)
        p_sub = sum(r["submitted"] for r in patwari_by_pct)
        patwari_totals = {
            "total": p_tot,
            "submitted": p_sub,
            "pct": _pct(p_sub, p_tot),
            "additions": sum(r["additions"] for r in patwari_by_pct) if has_additions else None,
        }
    else:
        patwari_totals = None

    # === Checker Wise ===
    # Columns: Checker | Sub-Division | Villages | Total Survey Nos | Submitted | Verified+Approved | % Completion | Additions
    # % Completion = (verified + approved + seek_clarification) / submitted * 100
    checker_rows = None
    checker_totals = None
    if "checker" in current_df.columns and current_df["checker"].str.strip().replace("", pd.NA).notna().any():
        c_df = current_df[current_df["checker"].str.strip() != ""].copy()
        c = (c_df.groupby("checker", as_index=False)
                 .agg(villages=("village", "count"),
                      villages_list=("village", lambda s: ", ".join(sorted(s.tolist()))),
                      subdivisions=("subdivision", lambda s: " / ".join(sorted(set(s)))),
                      total=("khasras", "sum"),
                      submitted=("submitted", "sum"),
                      verified=("verified", "sum"),
                      approved=("approved", "sum"),
                      seek_clarification=("seek_clarification", "sum"))
                 .sort_values("submitted", ascending=False))
        old_by_ck = from_df.groupby("checker")["submitted"].sum().to_dict() if has_additions and "checker" in from_df.columns else {}
        checker_rows = []
        for _, row in c.iterrows():
            v_plus_a = int(row["verified"]) + int(row["approved"])
            processed = v_plus_a + int(row["seek_clarification"])
            checker_rows.append({
                "name": row["checker"],
                "subdivision": row["subdivisions"],
                "villages_list": row["villages_list"],
                "villages": int(row["villages"]),
                "total": int(row["total"]),
                "submitted": int(row["submitted"]),
                "verified_plus_approved": v_plus_a,
                "processed_incl_seekclar": processed,
                "pct": _pct(processed, int(row["submitted"])),
                "additions": (int(row["submitted"]) - int(old_by_ck.get(row["checker"], 0))) if has_additions else None,
            })
        # Sort by % Completion ascending — strugglers first, best performers at bottom
        checker_rows.sort(key=lambda r: r["pct"])
        c_sub = sum(r["submitted"] for r in checker_rows)
        c_va = sum(r["verified_plus_approved"] for r in checker_rows)
        c_proc = sum(r["processed_incl_seekclar"] for r in checker_rows)
        checker_totals = {
            "villages": sum(r["villages"] for r in checker_rows),
            "total": sum(r["total"] for r in checker_rows),
            "submitted": c_sub,
            "verified_plus_approved": c_va,
            "pct": _pct(c_proc, c_sub),
            "additions": sum(r["additions"] for r in checker_rows) if has_additions else None,
        }

    # === Village Wise ===
    # Default: grouped tehsil-then-village. Filter/coloring: sorted by submitted desc + banding.
    if apply_coloring:
        if active_tehsils is not None:
            v_src = current_df[current_df["tehsil"].str.upper().isin(active_tehsils)]
        else:
            v_src = current_df
        v_df = v_src[["village", "tehsil", "patwari", "khasras", "submitted", "verified", "approved"]].copy()
        v_df = v_df.sort_values("submitted", ascending=False, kind="mergesort").reset_index(drop=True)
    else:
        v_df = current_df[["village", "tehsil", "patwari", "khasras", "submitted", "verified", "approved"]].copy()
        v_df = v_df.sort_values(["tehsil", "village"], kind="mergesort").reset_index(drop=True)
    old_by_village = from_df.groupby(["tehsil", "village"])["submitted"].sum().to_dict() if has_additions else {}
    village_rows = []
    for _, r in v_df.iterrows():
        additions = int(r["submitted"]) - int(old_by_village.get((r["tehsil"], r["village"]), 0)) if has_additions else None
        village_rows.append({
            "village": r["village"],
            "tehsil": r["tehsil"],
            "patwari": r["patwari"],
            "total": int(r["khasras"]),
            "submitted": int(r["submitted"]),
            "verified": int(r["verified"]) + int(r["approved"]),  # roll-in: verified+ = at-least-verified count
            "approved": int(r["approved"]),
            "additions": additions,
        })
    if apply_coloring:
        apply_bands(village_rows, "submitted")
    if village_rows:
        v_tot = sum(r["total"] for r in village_rows)
        v_sub = sum(r["submitted"] for r in village_rows)
        village_totals = {
            "total": v_tot,
            "submitted": v_sub,
            "verified": sum(r["verified"] for r in village_rows),
            "approved": sum(r["approved"] for r in village_rows),
            "pct": _pct(v_sub, v_tot),
            "additions": sum(r["additions"] for r in village_rows) if has_additions else None,
        }
    else:
        village_totals = None

    # Top cards / stats
    top_patwari = patwari_by_count[0] if patwari_by_count else None
    top_tehsil = tehsil_rows[0] if tehsil_rows else None
    grand = {
        "tehsils": int(len(t)),
        "villages": int(len(current_df)),
        "patwaris": int(current_df["patwari"].nunique()),
        "total_khasras": int(current_df["khasras"].sum()),
        "submitted": int(current_df["submitted"].sum()),
        "approved": int(current_df["approved"].sum()),
        "overall_pct": _pct(int(current_df["submitted"].sum()), int(current_df["khasras"].sum())),
        "overall_pct_approved": _pct(int(current_df["approved"].sum()), int(current_df["khasras"].sum())),
        "additions": (int(current_df["submitted"].sum()) - int(from_df["submitted"].sum())) if has_additions else None,
    }

    return {
        "tehsil_rows": tehsil_rows,
        "overall_rows": overall_rows,
        "overall_totals": overall_totals,
        "patwari_by_pct": patwari_by_pct,
        "patwari_by_count": patwari_by_count,
        "patwari_totals": patwari_totals,
        "checker_rows": checker_rows,
        "checker_totals": checker_totals,
        "village_rows": village_rows,
        "village_totals": village_totals,
        "top_patwari": top_patwari,
        "top_tehsil": top_tehsil,
        "grand": grand,
        "has_additions": has_additions,
    }


# ---------- Config ----------

def read_as_of_override():
    if os.path.exists("AS_OF.txt"):
        with open("AS_OF.txt", "r", encoding="utf-8") as f:
            text = f.read().strip()
        if text:
            return text
    return None


def format_date(d):
    return d.strftime("%d %b %Y")


def format_date_short(d):
    return d.strftime("%d %b")


def parse_date_param(s):
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None


def parse_tehsils_param(request_args):
    values = request_args.getlist("tehsils")
    if not values:
        return None
    cleaned = [v.strip() for v in values if v and v.strip()]
    if not cleaned:
        return None
    if any(v.lower() == "all" for v in cleaned):
        return "all"
    return set(v.upper() for v in cleaned)


def resolve_dates(snapshots, from_param, to_param):
    if not snapshots:
        return None, None
    dates_available = [d for d, _ in snapshots]
    to_date = parse_date_param(to_param) if to_param else None
    if not to_date or to_date not in dates_available:
        to_date = dates_available[0]
    from_date = parse_date_param(from_param) if from_param else None
    if not from_date or from_date not in dates_available or from_date >= to_date:
        older = [d for d in dates_available if d < to_date]
        from_date = older[0] if older else None
    return from_date, to_date


# ---------- Routes ----------

@app.route("/")
def landing():
    """Landing hub with two links to the sub-dashboards."""
    return render_template("landing.html")


@app.route("/data-punching")
def index():
    snapshots = all_snapshots()
    if not snapshots:
        return "No snapshots found. Add a file to snapshots/ folder.", 500

    # Mode: 'default' (current view) or 'target' (target based view)
    mode = request.args.get("mode", "default").strip().lower()
    if mode not in ("default", "target"):
        mode = "default"

    from_date, to_date = resolve_dates(snapshots, request.args.get("from"), request.args.get("to"))
    current_df = load_snapshot(snapshot_for_date(snapshots, to_date))
    from_df = load_snapshot(snapshot_for_date(snapshots, from_date)) if from_date else None
    tehsils_filter = parse_tehsils_param(request.args)
    views = build_views(current_df, from_df, tehsils_filter)

    # Only touch the plan file when the user actually wants the target view.
    # In default mode we do a cheap filesystem check so the toggle can render,
    # but we don't spend memory parsing the Excel.
    plan_available = has_plan_file()
    target_view = None
    baseline_info = None
    if mode == "target" and plan_available:
        try:
            plan_df = load_plan_file()
        except Exception as e:
            print(f"[target-view] Failed to load plan file: {e}", file=sys.stderr)
            plan_df = None
        if plan_df is not None:
            baseline = find_baseline_snapshot(snapshots, BASELINE_TARGET_DATE)
            baseline_df = None
            if baseline:
                b_date, b_path = baseline
                try:
                    baseline_df = load_snapshot(b_path)
                    baseline_info = {"date": b_date, "path": os.path.basename(b_path)}
                except Exception as e:
                    print(f"[target-view] Could not load baseline {b_path}: {e}", file=sys.stderr)
            try:
                target_view = build_target_view(current_df, baseline_df, plan_df)
            except Exception as e:
                print(f"[target-view] build_target_view failed: {e}", file=sys.stderr)
                target_view = None
            # Free heavy DataFrames so template rendering has more headroom
            del plan_df
            if baseline_df is not None:
                del baseline_df
            import gc; gc.collect()

    override = read_as_of_override()
    as_of_display = override if override else format_date(to_date)
    date_options = [(d.isoformat(), format_date(d)) for d, _ in snapshots]
    additions_label = f"Additions ({format_date_short(from_date)} → {format_date_short(to_date)})" if views["has_additions"] else None

    all_tehsils = sorted(current_df["tehsil"].dropna().unique().tolist())
    is_all_selected = tehsils_filter == "all"
    selected_tehsils = tehsils_filter if isinstance(tehsils_filter, set) else set()

    return render_template(
        "index.html",
        as_of=as_of_display,
        date_options=date_options,
        from_date_iso=from_date.isoformat() if from_date else None,
        to_date_iso=to_date.isoformat(),
        additions_label=additions_label,
        snapshot_count=len(snapshots),
        all_tehsils=all_tehsils,
        is_all_selected=is_all_selected,
        selected_tehsils=selected_tehsils,
        coloring_active=(tehsils_filter is not None),
        mode=mode,
        plan_available=plan_available,
        target_view=target_view,
        baseline_info=baseline_info,
        target_deadline=DEADLINE_DATE,
        **views,
    )


@app.route("/farmer-id")
def farmer_id():
    """Serve the farmer-ID generation HTML. User uploads it as templates/farmer_id.html
    via GitHub — each upload replaces the previous. If not yet uploaded, show a
    friendly placeholder. A non-intrusive "back to landing" link is injected at
    serve time so we never modify the user's uploaded file on disk."""
    path = os.path.join("templates", "farmer_id.html")
    if not os.path.exists(path):
        return render_template("farmer_id_missing.html"), 200
    with open(path, "r", encoding="utf-8") as f:
        content = f.read()

    # Inject a fixed-position back link right after the opening <body> tag. Use a
    # scoped <style> block with a specific ID so it can be responsive without
    # colliding with the uploaded file's own CSS.
    back_link = (
        '<style>'
        '#agr-back-hub{'
        'position:fixed;top:14px;left:14px;z-index:9999;'
        'background:#FDFBF5;color:#1F3F2E;'
        'padding:9px 16px;border:1px solid #D6D1C2;border-radius:6px;'
        'text-decoration:none;'
        "font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Arial,sans-serif;"
        'font-size:13px;font-weight:500;letter-spacing:0.02em;'
        'box-shadow:0 2px 8px rgba(0,0,0,0.12);'
        'transition:background 140ms,transform 140ms;'
        'display:inline-block;'
        '}'
        '#agr-back-hub:hover{background:#E8EEDE;transform:translateY(-1px);}'
        '@media (max-width:720px){'
        '#agr-back-hub{left:auto;right:14px;padding:7px 12px;font-size:12px;}'
        '}'
        '</style>'
        '<a href="/" id="agr-back-hub" '
        'aria-label="Back to Progress Monitoring landing page">'
        '\u2190 &nbsp;Progress Monitoring'
        '</a>'
    )
    body_re = re.compile(r"(<body[^>]*>)", re.IGNORECASE)
    m = body_re.search(content)
    if m:
        content = content[:m.end()] + back_link + content[m.end():]
    else:
        content = back_link + content

    return Response(content, mimetype="text/html; charset=utf-8")


def _check_download_password():
    expected = os.environ.get("DOWNLOAD_PASSWORD", "")
    if not expected:
        return False
    auth = request.authorization
    if not auth or auth.password != expected:
        return False
    return True


def _resolve_download_context():
    snapshots = all_snapshots()
    if not snapshots:
        return None
    from_date, to_date = resolve_dates(snapshots, request.args.get("from"), request.args.get("to"))
    current_df = load_snapshot(snapshot_for_date(snapshots, to_date))
    from_df = load_snapshot(snapshot_for_date(snapshots, from_date)) if from_date else None
    tehsils_filter = parse_tehsils_param(request.args)
    views = build_views(current_df, from_df, tehsils_filter)
    return {"views": views, "to_date": to_date, "from_date": from_date,
            "gap_days": (to_date - from_date).days if from_date else 0}


def _panels_for_view(views, view_key):
    """Return list of {title, type, key, rows, totals} for the view."""
    all_panels = [
        ("Tehsil Wise", "tehsil", views.get("tehsil_rows"), None, "tehsil"),
        ("Overall All", "overall", views.get("overall_rows"), views.get("overall_totals"), "overall-all"),
        ("Patwari Wise — By % Completion", "patwari", views.get("patwari_by_pct"), views.get("patwari_totals"), "patwari-pct"),
        ("Patwari Wise — By Total Submissions", "patwari", views.get("patwari_by_count"), views.get("patwari_totals"), "patwari-count"),
        ("Checker Wise", "checker", views.get("checker_rows"), views.get("checker_totals"), "checker"),
        ("Village Wise", "village", views.get("village_rows"), views.get("village_totals"), "village"),
    ]
    out = []
    for title, ptype, rows, totals, key in all_panels:
        if rows is None:
            continue
        if view_key != "all" and key != view_key:
            continue
        out.append({"title": title, "type": ptype, "rows": rows, "totals": totals, "key": key})
    return out


@app.route("/download.xlsx")
def download():
    if not _check_download_password():
        return Response("Authentication required", 401,
                        {"WWW-Authenticate": 'Basic realm="AgriStack Download"'})
    ctx = _resolve_download_context()
    if not ctx:
        return "No snapshots found.", 500
    view_key = request.args.get("view", "all")
    buf = _generate_workbook(ctx["views"], ctx["to_date"], ctx["from_date"], view_key)
    tail = "" if view_key == "all" else f"_{view_key}"
    filename = f"AGRISTACK_Dashboard_{ctx['to_date'].strftime('%Y-%m-%d')}{tail}.xlsx"
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)


@app.route("/download.pdf")
def download_pdf():
    if not _check_download_password():
        return Response("Authentication required", 401,
                        {"WWW-Authenticate": 'Basic realm="AgriStack Download"'})
    ctx = _resolve_download_context()
    if not ctx:
        return "No snapshots found.", 500
    view_key = request.args.get("view", "tehsil")
    panels = _panels_for_view(ctx["views"], view_key)
    if not panels:
        return "No data for this view.", 400

    additions_label = None
    if ctx["gap_days"] > 0:
        additions_label = f"Additions ({ctx['from_date'].strftime('%d %b')} → {ctx['to_date'].strftime('%d %b')})"

    # Landscape works for every view here (Overall, Tehsil, Patwari, Checker, Village are all wide)
    orientation = "landscape"

    html = render_template("print.html",
                           panels=panels,
                           as_of=format_date(ctx["to_date"]),
                           from_date_display=format_date(ctx["from_date"]) if ctx["from_date"] else None,
                           gap_days=ctx["gap_days"],
                           has_additions=ctx["views"]["has_additions"],
                           additions_label=additions_label,
                           grand=ctx["views"]["grand"],
                           orientation=orientation)
    buf = io.BytesIO()
    result = pisa.CreatePDF(html, dest=buf)
    if result.err:
        return "PDF generation failed", 500
    buf.seek(0)
    tail = "all" if view_key == "all" else view_key
    filename = f"AGRISTACK_Dashboard_{ctx['to_date'].strftime('%Y-%m-%d')}_{tail}.pdf"
    return send_file(buf, mimetype="application/pdf",
                     as_attachment=True, download_name=filename)


# ---------- Excel workbook ----------

def _generate_workbook(views, to_date, from_date, view_key="all"):
    def _include(key):
        return view_key == "all" or view_key == key

    hdr_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="1F3F2E", end_color="1F3F2E", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_font = Font(name="Arial", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    tot_font = Font(name="Arial", size=11, bold=True)
    tot_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    band_fills = {
        "green":  PatternFill(start_color="DDF0D8", end_color="DDF0D8", fill_type="solid"),
        "yellow": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        "red":    PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
    }

    def write_hdr(ws, headers):
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align; c.border = border
        ws.row_dimensions[1].height = 38

    def apply_band_fill(ws, row_num, num_cols, band):
        if not band or band not in band_fills:
            return
        for ci in range(1, num_cols + 1):
            ws.cell(row=row_num, column=ci).fill = band_fills[band]

    wb = Workbook()
    wb.remove(wb.active)
    grand = views["grand"]

    additions_hdr = None
    if views["has_additions"]:
        additions_hdr = f"Additions ({from_date.strftime('%d %b')} to {to_date.strftime('%d %b')})"

    # === TEHSIL WISE ===
    if _include("tehsil"):
        ws = wb.create_sheet("TEHSIL WISE")
        headers = ["S.NO", "TEHSIL", "NUMBER OF VILLAGES", "TOTAL SURVEY NOS",
                   "SUBMITTED", "% COMPLETION"]
        if additions_hdr:
            headers.append(additions_hdr)
        headers.append("% APPROVED")
        write_hdr(ws, headers)
        approved_col = len(headers)
        additions_col = len(headers) - 1 if additions_hdr else None
        for i, row in enumerate(views["tehsil_rows"], start=1):
            r = i + 1
            vals = [i, row["tehsil"], row["villages"], row["total"], row["submitted"], round(row["pct"], 2)]
            aligns = [center, left, center, center, center, center]
            if additions_hdr:
                vals.append(row["additions"] if row["additions"] is not None else "—")
                aligns.append(center)
            vals.append(round(row["pct_approved"], 2))
            aligns.append(center)
            for ci, (v, a) in enumerate(zip(vals, aligns), start=1):
                c = ws.cell(row=r, column=ci, value=v)
                c.alignment = a; c.font = body_font; c.border = border
                if ci in (3, 4, 5): c.number_format = "#,##0"
                if ci == 6: c.number_format = '0.00"%"'
                if additions_col and ci == additions_col and isinstance(v, int):
                    c.number_format = "+#,##0;-#,##0;0"
                if ci == approved_col: c.number_format = '0.00"%"'
        # TOTAL
        tt = len(views["tehsil_rows"]) + 2
        ws.cell(row=tt, column=2, value="TOTAL").alignment = left
        ws.cell(row=tt, column=3, value=grand["villages"]).alignment = center
        ws.cell(row=tt, column=4, value=grand["total_khasras"]).alignment = center
        ws.cell(row=tt, column=5, value=grand["submitted"]).alignment = center
        ws.cell(row=tt, column=6, value=round(grand["overall_pct"], 2)).alignment = center
        if additions_col:
            ws.cell(row=tt, column=additions_col, value=grand["additions"] if grand["additions"] is not None else "—").alignment = center
        ws.cell(row=tt, column=approved_col, value=round(grand["overall_pct_approved"], 2)).alignment = center
        for c in range(1, len(headers) + 1):
            cc = ws.cell(row=tt, column=c)
            cc.font = tot_font; cc.fill = tot_fill; cc.border = border
            if c in (3, 4, 5): cc.number_format = "#,##0"
            if c == 6: cc.number_format = '0.00"%"'
            if additions_col and c == additions_col and isinstance(cc.value, int):
                cc.number_format = "+#,##0;-#,##0;0"
            if c == approved_col: cc.number_format = '0.00"%"'
        widths = [7, 16, 16, 18, 14, 14]
        if additions_hdr: widths.append(22)
        widths.append(14)
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    # === OVERALL ALL ===
    if _include("overall-all"):
        ws = wb.create_sheet("OVERALL ALL")
        headers = ["S.NO", "TEHSIL", "TOTAL KHASRAS", "KHASRAS SUBMITTED",
                   "KHASRAS VERIFIED", "KHASRAS APPROVED"]
        write_hdr(ws, headers)
        for i, row in enumerate(views["overall_rows"], start=1):
            r = i + 1
            vals = [i, row["tehsil"], row["total"], row["submitted"], row["verified"], row["approved"]]
            aligns = [center, left, center, center, center, center]
            for ci, (v, a) in enumerate(zip(vals, aligns), start=1):
                c = ws.cell(row=r, column=ci, value=v)
                c.alignment = a; c.font = body_font; c.border = border
                if ci in (3, 4, 5, 6): c.number_format = "#,##0"
        ot = views["overall_totals"]
        tt = len(views["overall_rows"]) + 2
        ws.cell(row=tt, column=2, value="TOTAL").alignment = left
        ws.cell(row=tt, column=3, value=ot["total"]).alignment = center
        ws.cell(row=tt, column=4, value=ot["submitted"]).alignment = center
        ws.cell(row=tt, column=5, value=ot["verified"]).alignment = center
        ws.cell(row=tt, column=6, value=ot["approved"]).alignment = center
        for c in range(1, 7):
            cc = ws.cell(row=tt, column=c)
            cc.font = tot_font; cc.fill = tot_fill; cc.border = border
            if c in (3, 4, 5, 6): cc.number_format = "#,##0"
        for i, w in enumerate([7, 16, 18, 20, 20, 20], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    # === Patwari sheets ===
    for title, rows, key in [("PATWARI BY %", views["patwari_by_pct"], "patwari-pct"),
                              ("PATWARI BY COUNT", views["patwari_by_count"], "patwari-count")]:
        if not _include(key):
            continue
        show_effort = (key == "patwari-pct")
        ws = wb.create_sheet(title)
        headers = ["S.NO", "NAME OF PATWARI", "TEHSIL", "VILLAGES", "TOTAL SURVEY NOS",
                   "SUBMITTED", "% COMPLETION"]
        if show_effort:
            headers.append("RELATIVE EFFORT")
        if additions_hdr:
            headers.append(additions_hdr)
        write_hdr(ws, headers)
        effort_col = 8 if show_effort else None
        additions_col = len(headers) if additions_hdr else None
        total_cols = len(headers)
        for i, row in enumerate(rows, start=1):
            r = i + 1
            vals = [i, row["patwari"], row["tehsil"], row["villages_list"], row["total"],
                    row["submitted"], round(row["pct"], 2)]
            aligns = [center, left, left, left, center, center, center]
            if show_effort:
                re_v = row.get("relative_effort")
                vals.append(round(re_v, 2) if re_v is not None else "—")
                aligns.append(center)
            if additions_hdr:
                vals.append(row["additions"] if row["additions"] is not None else "—")
                aligns.append(center)
            for ci, (v, a) in enumerate(zip(vals, aligns), start=1):
                c = ws.cell(row=r, column=ci, value=v)
                c.alignment = a; c.font = body_font; c.border = border
                if ci in (5, 6): c.number_format = "#,##0"
                if ci == 7: c.number_format = '0.00"%"'
                if effort_col and ci == effort_col and isinstance(v, (int, float)):
                    c.number_format = '0.00"%"'
                if additions_col and ci == additions_col and isinstance(v, int):
                    c.number_format = "+#,##0;-#,##0;0"
            apply_band_fill(ws, r, total_cols, row.get("band"))
        # TOTAL
        pt = len(rows) + 2
        pt_tot = views.get("patwari_totals")
        use_filtered = pt_tot is not None and any(r.get("band") for r in rows)
        ws.cell(row=pt, column=2, value=("TOTAL (Filtered)" if use_filtered else "TOTAL")).alignment = left
        if use_filtered:
            ws.cell(row=pt, column=5, value=pt_tot["total"]).alignment = center
            ws.cell(row=pt, column=6, value=pt_tot["submitted"]).alignment = center
            ws.cell(row=pt, column=7, value=round(pt_tot["pct"], 2)).alignment = center
            if additions_col:
                ws.cell(row=pt, column=additions_col, value=pt_tot["additions"] if pt_tot["additions"] is not None else "—").alignment = center
        else:
            ws.cell(row=pt, column=5, value=grand["total_khasras"]).alignment = center
            ws.cell(row=pt, column=6, value=grand["submitted"]).alignment = center
            ws.cell(row=pt, column=7, value=round(grand["overall_pct"], 2)).alignment = center
            if additions_col:
                ws.cell(row=pt, column=additions_col, value=grand["additions"] if grand["additions"] is not None else "—").alignment = center
        for c in range(1, total_cols + 1):
            cc = ws.cell(row=pt, column=c)
            cc.font = tot_font; cc.fill = tot_fill; cc.border = border
            if c in (5, 6): cc.number_format = "#,##0"
            if c == 7: cc.number_format = '0.00"%"'
            if show_effort and c == effort_col: cc.number_format = '0.00"%"'
            if additions_col and c == additions_col and isinstance(cc.value, int):
                cc.number_format = "+#,##0;-#,##0;0"
        widths = [7, 26, 14, 50, 18, 14, 14]
        if show_effort: widths.append(16)
        if additions_hdr: widths.append(22)
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    # === CHECKER WISE ===
    if _include("checker") and views.get("checker_rows"):
        ws = wb.create_sheet("CHECKER WISE")
        headers = ["S.NO", "CHECKER", "SUB-DIVISION", "VILLAGES", "TOTAL SURVEY NOS",
                   "SUBMITTED", "VERIFIED + APPROVED", "% COMPLETION"]
        if additions_hdr:
            headers.append(additions_hdr)
        write_hdr(ws, headers)
        additions_col = len(headers) if additions_hdr else None
        # Formula note as row 2 (comment above table)
        note_row = 2
        ws.cell(row=note_row, column=1, value="% Completion = (Verified + Approved + Seek Clarification) ÷ Submitted × 100").alignment = left
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=len(headers))
        ws.cell(row=note_row, column=1).font = Font(name="Arial", size=10, italic=True, color="4A6B4E")
        data_start = 3
        for i, row in enumerate(views["checker_rows"], start=1):
            r = i + data_start - 1
            vals = [i, row["name"], row["subdivision"], row["villages_list"],
                    row["total"], row["submitted"], row["verified_plus_approved"], round(row["pct"], 2)]
            aligns = [center, left, left, left, center, center, center, center]
            if additions_hdr:
                vals.append(row["additions"] if row["additions"] is not None else "—")
                aligns.append(center)
            for ci, (v, a) in enumerate(zip(vals, aligns), start=1):
                c = ws.cell(row=r, column=ci, value=v)
                c.alignment = a; c.font = body_font; c.border = border
                if ci in (5, 6, 7): c.number_format = "#,##0"
                if ci == 8: c.number_format = '0.00"%"'
                if additions_col and ci == additions_col and isinstance(v, int):
                    c.number_format = "+#,##0;-#,##0;0"
        ct = views.get("checker_totals")
        if ct:
            pt = data_start + len(views["checker_rows"])
            ws.cell(row=pt, column=2, value="TOTAL").alignment = left
            ws.cell(row=pt, column=5, value=ct["total"]).alignment = center
            ws.cell(row=pt, column=6, value=ct["submitted"]).alignment = center
            ws.cell(row=pt, column=7, value=ct["verified_plus_approved"]).alignment = center
            ws.cell(row=pt, column=8, value=round(ct["pct"], 2)).alignment = center
            if additions_col:
                ws.cell(row=pt, column=additions_col, value=ct["additions"] if ct["additions"] is not None else "—").alignment = center
            for c in range(1, len(headers) + 1):
                cc = ws.cell(row=pt, column=c)
                cc.font = tot_font; cc.fill = tot_fill; cc.border = border
                if c in (5, 6, 7): cc.number_format = "#,##0"
                if c == 8: cc.number_format = '0.00"%"'
                if additions_col and c == additions_col and isinstance(cc.value, int):
                    cc.number_format = "+#,##0;-#,##0;0"
        widths = [7, 32, 14, 46, 16, 14, 18, 14]
        if additions_hdr: widths.append(22)
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = f"A{data_start}"

    # === VILLAGE WISE ===
    if _include("village"):
        ws = wb.create_sheet("VILLAGE WISE")
        headers = ["S.NO", "TEHSIL", "VILLAGE", "TOTAL SURVEY NOS", "NAME OF PATWARI",
                   "SUBMITTED", "VERIFIED", "APPROVED"]
        if additions_hdr:
            headers.append(additions_hdr)
        write_hdr(ws, headers)
        additions_col = len(headers) if additions_hdr else None
        total_cols = len(headers)
        for i, row in enumerate(views["village_rows"], start=1):
            r = i + 1
            vals = [i, row["tehsil"], row["village"], row["total"], row["patwari"],
                    row["submitted"], row["verified"], row["approved"]]
            aligns = [center, left, left, center, left, center, center, center]
            if additions_hdr:
                vals.append(row["additions"] if row["additions"] is not None else "—")
                aligns.append(center)
            for ci, (v, a) in enumerate(zip(vals, aligns), start=1):
                c = ws.cell(row=r, column=ci, value=v)
                c.alignment = a; c.font = body_font; c.border = border
                if ci in (4, 6, 7, 8): c.number_format = "#,##0"
                if additions_col and ci == additions_col and isinstance(v, int):
                    c.number_format = "+#,##0;-#,##0;0"
            apply_band_fill(ws, r, total_cols, row.get("band"))
        vt = len(views["village_rows"]) + 2
        v_tot = views.get("village_totals")
        use_filtered = v_tot is not None and any(r.get("band") for r in views["village_rows"])
        ws.cell(row=vt, column=2, value=("TOTAL (Filtered)" if use_filtered else "TOTAL")).alignment = left
        if use_filtered:
            ws.cell(row=vt, column=4, value=v_tot["total"]).alignment = center
            ws.cell(row=vt, column=6, value=v_tot["submitted"]).alignment = center
            ws.cell(row=vt, column=7, value=v_tot["verified"]).alignment = center
            ws.cell(row=vt, column=8, value=v_tot["approved"]).alignment = center
            if additions_col:
                ws.cell(row=vt, column=additions_col, value=v_tot["additions"] if v_tot["additions"] is not None else "—").alignment = center
        else:
            ws.cell(row=vt, column=4, value=grand["total_khasras"]).alignment = center
            ws.cell(row=vt, column=6, value=grand["submitted"]).alignment = center
            ov = views["overall_totals"]
            ws.cell(row=vt, column=7, value=ov["verified"]).alignment = center
            ws.cell(row=vt, column=8, value=ov["approved"]).alignment = center
            if additions_col:
                ws.cell(row=vt, column=additions_col, value=grand["additions"] if grand["additions"] is not None else "—").alignment = center
        for c in range(1, total_cols + 1):
            cc = ws.cell(row=vt, column=c)
            cc.font = tot_font; cc.fill = tot_fill; cc.border = border
            if c in (4, 6, 7, 8): cc.number_format = "#,##0"
            if additions_col and c == additions_col and isinstance(cc.value, int):
                cc.number_format = "+#,##0;-#,##0;0"
        widths = [7, 14, 28, 20, 28, 14, 14, 14]
        if additions_hdr: widths.append(22)
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    # === META (always) ===
    ws4 = wb.create_sheet("META")
    ws4.cell(row=1, column=1, value="Data as of").font = tot_font
    ws4.cell(row=1, column=2, value=to_date.strftime("%d %b %Y"))
    ws4.cell(row=2, column=1, value="Additions from").font = tot_font
    ws4.cell(row=2, column=2, value=from_date.strftime("%d %b %Y") if from_date else "—")
    ws4.column_dimensions["A"].width = 18
    ws4.column_dimensions["B"].width = 20

    if len(wb.sheetnames) == 1:
        info = wb.create_sheet("INFO", 0)
        info.cell(row=1, column=1, value=f"No data for view: {view_key}")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.route("/healthz")
def healthz():
    return "ok", 200


@app.route("/debug-files")
def debug_files():
    """Diagnostic: shows exactly what files the running app sees in the
    filesystem — the same information has_plan_file() uses. Handy when the
    Target Based view toggle isn't appearing."""
    lines = []
    lines.append(f"Working directory: {os.getcwd()}")
    lines.append(f"SNAPSHOTS_DIR: {SNAPSHOTS_DIR}")
    lines.append(f"SNAPSHOTS_DIR exists: {os.path.isdir(SNAPSHOTS_DIR)}")
    lines.append("")
    lines.append(f"reference.xlsx present in repo root: {os.path.exists(REFERENCE_PATH)}")
    lines.append("")
    lines.append("Contents of snapshots/ folder:")
    if os.path.isdir(SNAPSHOTS_DIR):
        for name in sorted(os.listdir(SNAPSHOTS_DIR)):
            path = os.path.join(SNAPSHOTS_DIR, name)
            size = os.path.getsize(path) if os.path.isfile(path) else "(dir)"
            matches_plan = bool(re.match(r"plan[_ ]\d{4}-\d{2}-\d{2}\.xlsx$", name, re.I))
            lines.append(f"  {name!r}  ({size} bytes)  plan-match: {matches_plan}")
    else:
        lines.append("  (directory does not exist)")
    lines.append("")
    lines.append(f"has_plan_file() returns: {has_plan_file()}")
    lines.append(f"all_snapshots() returns: {[(d.isoformat(), os.path.basename(p)) for d, p in all_snapshots()]}")
    return "<pre style='font-family:monospace;font-size:13px;padding:20px;background:#f6f3ec;color:#1f3f2e;line-height:1.5'>" + "\n".join(lines) + "</pre>"


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
